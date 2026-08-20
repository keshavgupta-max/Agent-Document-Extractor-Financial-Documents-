"""Excel Document Parser using openpyxl."""

import io
import openpyxl
from logger import logger
from tools.parser.exceptions import CorruptedDocument, ParserExecutionError
from tools.parser.models import PageContent, ParsedDocument, TableContent


class ExcelParser:
    """Parser dedicated to extracting worksheets, rows, and tables from Excel (.xlsx, .xls) spreadsheets."""

    def parse(
        self,
        file_bytes: bytes,
        document_id: str,
        storage_path: str,
        mime_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ) -> ParsedDocument:
        """Extracts worksheet text representations and tabular data matrices from Excel bytes.

        Raises:
            CorruptedDocument: If the Excel workbook cannot be loaded.
            ParserExecutionError: If an error occurs during spreadsheet parsing.
        """
        try:
            workbook_stream = io.BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(workbook_stream, data_only=True)
        except Exception as exc:
            error_msg = f"Failed to open Excel workbook (corrupted or invalid): {str(exc)}"
            logger.error(error_msg)
            raise CorruptedDocument(error_msg) from exc

        pages = []
        tables = []

        try:
            sheet_names = workbook.sheetnames
            for page_idx, sheet_name in enumerate(sheet_names, start=1):
                sheet = workbook[sheet_name]
                sheet_rows = []

                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    formatted_row = [
                        str(cell).strip() if cell is not None else "" for cell in row
                    ]
                    if any(formatted_row):
                        sheet_rows.append(formatted_row)

                if not sheet_rows:
                    continue

                # Construct textual summary of worksheet for RAG/full-text query
                text_lines = [f"--- Worksheet: {sheet_name} ---"]
                for row in sheet_rows:
                    text_lines.append(" | ".join(row))
                sheet_text = "\n".join(text_lines)

                pages.append(PageContent(page_number=page_idx, text=sheet_text))

                # Treat sheet data as a TableContent object
                headers = sheet_rows[0]
                rows = sheet_rows[1:] if len(sheet_rows) > 1 else []
                tables.append(
                    TableContent(
                        table_index=page_idx - 1,
                        page_number=page_idx,
                        headers=headers,
                        rows=rows,
                    )
                )

            metadata = {
                "sheet_names": sheet_names,
                "total_sheets": len(sheet_names),
            }

            workbook.close()

            return ParsedDocument(
                document_id=document_id,
                storage_path=storage_path,
                file_extension=".xlsx",
                mime_type=mime_type,
                page_count=len(pages),
                pages=pages,
                tables=tables,
                metadata=metadata,
                parsing_status="SUCCESS",
            )

        except Exception as exc:
            workbook.close()
            error_msg = f"Error parsing Excel spreadsheet for document {document_id}: {str(exc)}"
            logger.error(error_msg, exc_info=True)
            raise ParserExecutionError(error_msg) from exc